"""
Template Writer Module
======================
Handles writing datasets to Excel template sheets and managing unused sheets.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
from datetime import date, datetime
from typing import Dict, List, Optional, Any, Callable
from dataclasses import dataclass


@dataclass
class SheetConfig:
    """Configuration for writing to a specific template sheet."""
    data_start_row: int = 7
    data_start_col: int = 1
    cell_user_input: tuple = (1, 2)      # B1 - Data Pull Name
    cell_services: tuple = (2, 2)        # B2 - Services
    cell_distributors: tuple = (3, 2)    # B3 - Distributors
    cell_footnotes: tuple = (4, 2)       # B4 - Footnotes
    cell_date: tuple = (5, 2)            # B5 - Date


class TemplateWriter:
    """
    Handles writing multiple datasets to an Excel template.
    Manages sheet assignments and deletion of unused sheets.
    """

    def __init__(self, template_path: str, config: SheetConfig = None):
        """
        Initialize the template writer.

        Args:
            template_path: Path to the Excel template file
            config: Sheet configuration (uses defaults if not provided)
        """
        self.template_path = template_path
        self.config = config or SheetConfig()
        self._wb = None
        self._template_sheets = []

    def load_template(self) -> List[str]:
        """Load the template and return available sheet names."""
        self._wb = load_workbook(self.template_path)
        self._template_sheets = self._wb.sheetnames.copy()
        return self._template_sheets

    @property
    def template_sheets(self) -> List[str]:
        """Get list of template sheet names."""
        return self._template_sheets

    def write_dataset_to_sheet(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        data_pull_name: str = "",
        selected_footnotes: List[str] = None,
        footnotes_table: Dict[str, str] = None,
        convert_numeric_func: Callable = None,
        try_parse_number_func: Callable = None,
        try_parse_date_func: Callable = None,
        is_date_column_func: Callable = None,
        is_rate_column_func: Callable = None,
        format_column_name_func: Callable = None,
        extract_unique_values_func: Callable = None
    ):
        """
        Write a dataset to a specific template sheet.

        This method replicates the logic from the original write_to_template function
        but targets a specific sheet.
        """
        if self._wb is None:
            raise RuntimeError("Template not loaded. Call load_template() first.")

        if sheet_name not in self._wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in template.")

        ws = self._wb[sheet_name]

        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')

        cfg = self.config

        # B1: Data Pull Name
        ws.cell(row=cfg.cell_user_input[0], column=cfg.cell_user_input[1]).value = data_pull_name

        # B2: Services
        if extract_unique_values_func:
            services = extract_unique_values_func(df, [col for col in df.columns if "service" in col.lower()])
            ws.cell(row=cfg.cell_services[0], column=cfg.cell_services[1]).value = services

        # B3: Distributors
        if extract_unique_values_func:
            distributors = extract_unique_values_func(df, [col for col in df.columns if "distributor" in col.lower()])
            ws.cell(row=cfg.cell_distributors[0], column=cfg.cell_distributors[1]).value = distributors

        # B4: Selected footnotes
        if selected_footnotes and footnotes_table:
            footnote_text = "\n\n".join([f"{footnotes_table[metric]}" for metric in selected_footnotes if metric in footnotes_table])
            cell_b4 = ws.cell(row=cfg.cell_footnotes[0], column=cfg.cell_footnotes[1])
            cell_b4.value = footnote_text
            cell_b4.alignment = Alignment(wrap_text=True, vertical='top')

        # B5: Date
        date_cell = ws.cell(row=cfg.cell_date[0], column=cfg.cell_date[1])
        date_cell.value = date.today()
        date_cell.number_format = 'yyyy-mm-dd'

        # Convert numbers before writing
        df_converted = convert_numeric_func(df) if convert_numeric_func else df.copy()

        # Identify column types
        date_cols = []
        rate_cols = []
        numeric_cols = []

        if is_date_column_func:
            date_cols = [col for col in df.columns if is_date_column_func(df[col])]

        if is_rate_column_func:
            rate_cols = [col for col in df.columns if is_rate_column_func(col)]

        # Detect numeric columns
        for col in df_converted.columns:
            if col not in date_cols:
                if try_parse_number_func:
                    num_test = df_converted[col].apply(
                        lambda x: try_parse_number_func(x, preserve_percent=True)[0]
                        if try_parse_number_func(x, preserve_percent=True) else None
                    )
                    if num_test.notna().sum() > len(df_converted[col].dropna()) * 0.5:
                        numeric_cols.append(col)

        right_aligned_cols = set(rate_cols) | set(numeric_cols)

        # Write column headers
        for col_idx, col_name in enumerate(df.columns):
            cell = ws.cell(row=cfg.data_start_row, column=cfg.data_start_col + col_idx)
            cell.value = format_column_name_func(col_name) if format_column_name_func else col_name
            if col_name in right_aligned_cols:
                cell.alignment = right_align
            else:
                cell.alignment = left_align

        # Write data rows
        for row_idx, row in enumerate(df_converted.itertuples(index=False), start=1):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=cfg.data_start_row + row_idx, column=cfg.data_start_col + col_idx)
                col_name = df.columns[col_idx]

                if col_name in date_cols:
                    if try_parse_date_func:
                        parsed_date = try_parse_date_func(value)
                        if parsed_date:
                            cell.value = parsed_date
                            cell.number_format = 'MMM-YY'
                        else:
                            cell.value = value if pd.notna(value) else ""
                    else:
                        cell.value = value if pd.notna(value) else ""
                    cell.alignment = left_align

                elif col_name in rate_cols:
                    if try_parse_number_func:
                        result = try_parse_number_func(value, preserve_percent=True)
                        num, was_percent = result if result else (None, False)
                        if num is not None:
                            cell.value = num
                            cell.number_format = '0.00%'
                        else:
                            cell.value = value if pd.notna(value) else ""
                    else:
                        cell.value = value if pd.notna(value) else ""
                    cell.alignment = right_align

                elif col_name in numeric_cols:
                    if try_parse_number_func:
                        result = try_parse_number_func(value, preserve_percent=True)
                        num, was_percent = result if result else (None, False)
                        if num is not None:
                            if was_percent:
                                cell.value = num
                                cell.number_format = '0.00%'
                            else:
                                cell.value = num
                                cell.number_format = '#,##0'
                        else:
                            cell.value = value if pd.notna(value) else ""
                    else:
                        cell.value = value if pd.notna(value) else ""
                    cell.alignment = right_align

                else:
                    cell.value = value if pd.notna(value) else ""
                    cell.alignment = left_align

        # Delete columns marked "Delete"
        cols_to_delete = []
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=cfg.data_start_row, column=col_idx).value
            if header_value and str(header_value).strip().lower() == "delete":
                cols_to_delete.append(col_idx)

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

    def delete_unused_sheets(self, used_sheets: List[str]):
        """
        Delete template sheets that are not in the used_sheets list.

        Args:
            used_sheets: List of sheet names that have data assigned
        """
        if self._wb is None:
            raise RuntimeError("Template not loaded. Call load_template() first.")

        sheets_to_delete = [s for s in self._template_sheets if s not in used_sheets]

        for sheet_name in sheets_to_delete:
            if sheet_name in self._wb.sheetnames:
                del self._wb[sheet_name]

    def save_to_buffer(self) -> BytesIO:
        """Save the workbook to a BytesIO buffer and return it."""
        if self._wb is None:
            raise RuntimeError("Template not loaded. Call load_template() first.")

        buffer = BytesIO()
        self._wb.save(buffer)
        buffer.seek(0)
        return buffer

    def close(self):
        """Close the workbook."""
        if self._wb:
            self._wb.close()
            self._wb = None


def write_multi_dataset_template(
    template_path: str,
    dataset_sheet_mapping: Dict[int, Any],  # {dataset_id: (sheet_name, df, footnotes, data_pull_name)}
    footnotes_table: Dict[str, str],
    helper_funcs: Dict[str, Callable]
) -> Optional[BytesIO]:
    """
    Write multiple datasets to a template with custom sheet names.
    Dataset 1 goes to first sheet, Dataset 2 to second sheet, etc.
    Unused template sheets are deleted.

    Args:
        template_path: Path to the Excel template
        dataset_sheet_mapping: Dict mapping dataset_id to (sheet_name, df, footnotes, data_pull_name) tuples
            Datasets are ordered by dataset_id (1, 2, 3...)
        footnotes_table: Dict of metric names to footnote text
        helper_funcs: Dict of helper functions from main app:
            - convert_numeric_columns
            - try_parse_number
            - try_parse_date
            - is_date_column
            - is_rate_column
            - format_column_name
            - extract_unique_values

    Returns:
        BytesIO buffer containing the output Excel file, or None on error
    """
    try:
        writer = TemplateWriter(template_path)
        template_sheets = writer.load_template()

        used_template_sheets = []

        # Process datasets in order by dataset_id
        sorted_dataset_ids = sorted(dataset_sheet_mapping.keys())

        for idx, dataset_id in enumerate(sorted_dataset_ids):
            dataset_info = dataset_sheet_mapping[dataset_id]

            # Unpack the tuple: (sheet_name, df, footnotes, data_pull_name)
            target_sheet_name, df, footnotes, data_pull_name = dataset_info

            # Use the corresponding template sheet (by position)
            if idx < len(template_sheets):
                template_sheet = template_sheets[idx]

                # Rename the template sheet to the user's custom name
                if template_sheet in writer._wb.sheetnames:
                    ws = writer._wb[template_sheet]
                    ws.title = target_sheet_name

                used_template_sheets.append(target_sheet_name)

                writer.write_dataset_to_sheet(
                    sheet_name=target_sheet_name,
                    df=df,
                    data_pull_name=data_pull_name,
                    selected_footnotes=footnotes,
                    footnotes_table=footnotes_table,
                    convert_numeric_func=helper_funcs.get('convert_numeric_columns'),
                    try_parse_number_func=helper_funcs.get('try_parse_number'),
                    try_parse_date_func=helper_funcs.get('try_parse_date'),
                    is_date_column_func=helper_funcs.get('is_date_column'),
                    is_rate_column_func=helper_funcs.get('is_rate_column'),
                    format_column_name_func=helper_funcs.get('format_column_name'),
                    extract_unique_values_func=helper_funcs.get('extract_unique_values')
                )
            else:
                print(f"Warning: Not enough template sheets for dataset {dataset_id}, skipping.")

        # Delete unused sheets (sheets beyond what we need)
        writer.delete_unused_sheets(used_template_sheets)

        buffer = writer.save_to_buffer()
        writer.close()

        return buffer

    except Exception as e:
        print(f"Error writing multi-dataset template: {e}")
        import traceback
        traceback.print_exc()
        return None

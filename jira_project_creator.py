#!/usr/bin/env python3
"""
Jira Project Creator
Creates a project folder and working Excel file from a Jira ticket URL.

Double-click to run or: python jira_project_creator.py
"""

import requests
import re
import logging
import os
import subprocess
import platform
from requests.auth import HTTPBasicAuth
from pathlib import Path
from dotenv import load_dotenv
import tkinter as tk
from tkinter import simpledialog, filedialog, messagebox
from openpyxl import Workbook

# Load environment variables from .env file
load_dotenv(Path(__file__).parent / ".env")

# =============================================================================
# CONFIGURATION
# =============================================================================

class Config:
    """Configuration settings loaded from .env file."""
    JIRA_BASE_URL = os.environ.get("JIRA_BASE_URL", "").strip().strip('"')
    EMAIL = os.environ.get("JIRA_EMAIL", "").strip().strip('"')
    API_TOKEN = os.environ.get("JIRA_API_TOKEN", "").strip().strip('"')
    DEFAULT_SAVE_LOCATION = Path(__file__).parent  # Save in level-up folder by default

    @classmethod
    def validate(cls):
        """Validate that all required configuration is present."""
        missing = []
        if not cls.API_TOKEN:
            missing.append("JIRA_API_TOKEN")
        if not cls.EMAIL:
            missing.append("JIRA_EMAIL")
        if not cls.JIRA_BASE_URL:
            missing.append("JIRA_BASE_URL")

        if missing:
            raise ValueError(f"Missing required environment variables in .env: {', '.join(missing)}")


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging():
    """Set up logging configuration."""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler()]
    )
    return logging.getLogger(__name__)


# =============================================================================
# JIRA API
# =============================================================================

class JiraAPI:
    """Handle Jira API interactions."""

    def __init__(self, base_url, email, api_token):
        self.base_url = base_url.rstrip("/")
        self.auth = HTTPBasicAuth(email, api_token)
        self.headers = {"Accept": "application/json"}
        self.logger = logging.getLogger(__name__)

    def extract_ticket_key(self, url):
        """Extract ticket key from Jira URL."""
        match = re.search(r"([A-Z][A-Z0-9]+-\d+)", url)
        if match:
            return match.group(1)
        return None

    def get_ticket_info(self, ticket_key):
        """Fetch ticket information from Jira API."""
        api_url = f"{self.base_url}/rest/api/3/issue/{ticket_key}"

        try:
            response = requests.get(api_url, headers=self.headers, auth=self.auth, timeout=10)
            self.logger.info(f"API response status: {response.status_code}")

            if response.status_code == 200:
                data = response.json()
                summary = data["fields"]["summary"]
                status = data["fields"]["status"]["name"]
                assignee = data["fields"].get("assignee")
                assignee_name = assignee.get("displayName") if assignee else "Unassigned"

                # Extract description (handle Atlassian Document Format)
                description = data["fields"].get("description")
                if description and isinstance(description, dict):
                    description = self._extract_adf_text(description)

                return {
                    "key": ticket_key,
                    "summary": summary,
                    "status": status,
                    "assignee": assignee_name,
                    "description": description or "No description provided."
                }
            else:
                self.logger.error(f"API Error {response.status_code}: {response.text}")
                return None

        except requests.exceptions.RequestException as e:
            self.logger.error(f"Network error: {e}")
            return None

    def _extract_adf_text(self, adf):
        """Extract plain text from Atlassian Document Format."""
        texts = []

        def recurse(node):
            if isinstance(node, dict):
                if node.get("type") == "text":
                    texts.append(node.get("text", ""))
                for child in node.get("content", []):
                    recurse(child)
            elif isinstance(node, list):
                for item in node:
                    recurse(item)

        recurse(adf)
        return "\n".join(texts)


# =============================================================================
# PROJECT CREATOR
# =============================================================================

class ProjectCreator:
    """Handle project folder and file creation."""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def sanitize_name(self, name):
        """Convert a string to a valid file/folder name."""
        # Remove illegal characters
        sanitized = re.sub(r'[<>"/\\|?*\[\]:]+', ' ', name)
        # Remove consecutive spaces
        sanitized = re.sub(r' +', ' ', sanitized)
        # Strip and limit length
        return sanitized.strip()[:100]

    def create_project(self, base_path, ticket, jira_url):
        """Create project folder and working Excel file."""
        # Folder name: PROJ-123 | Ticket Title
        folder_name = f"{ticket['key']} | {self.sanitize_name(ticket['summary'])}"
        folder_path = Path(base_path) / folder_name

        try:
            folder_path.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"Folder created: {folder_path}")

            # Create working Excel file: WORKING - PROJ-123 - Ticket Title.xlsx
            file_name = f"WORKING - {ticket['key']} - {self.sanitize_name(ticket['summary'])}.xlsx"
            file_path = folder_path / file_name

            # Create blank Excel with Jira link in A1
            wb = Workbook()
            ws = wb.active
            ws["A1"] = jira_url
            ws["A1"].hyperlink = jira_url
            ws["A1"].style = "Hyperlink"
            wb.save(file_path)

            self.logger.info(f"Working file created: {file_path}")

            return folder_path, file_path

        except Exception as e:
            self.logger.error(f"Error creating project: {e}")
            return None, None

    def open_folder(self, folder_path):
        """Open folder in file explorer."""
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(folder_path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", folder_path])
            else:  # Linux
                subprocess.run(["xdg-open", folder_path])
            self.logger.info(f"Opened folder: {folder_path}")
        except Exception as e:
            self.logger.error(f"Error opening folder: {e}")


# =============================================================================
# USER INTERFACE
# =============================================================================

class UserInterface:
    """Handle user input through GUI dialogs."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.withdraw()  # Hide main window
        self.logger = logging.getLogger(__name__)

    def get_jira_url(self):
        """Get Jira URL from user."""
        jira_url = simpledialog.askstring(
            "Jira URL",
            "Enter Jira issue URL:"
        )

        if not jira_url:
            self.logger.info("No URL entered by user")
            return None

        return jira_url.strip()

    def get_save_location(self, default_path=None):
        """Get save location from user."""
        initial_dir = str(default_path) if default_path and default_path.exists() else None

        folder_selected = filedialog.askdirectory(
            title="Choose where to save the project folder",
            initialdir=initial_dir
        )

        if not folder_selected:
            self.logger.info("No folder selected by user")
            return None

        return folder_selected

    def ask_open_folder(self):
        """Ask user if they want to open the destination folder."""
        return messagebox.askyesno(
            "Open Folder?",
            "Would you like to open the project folder?"
        )

    def show_success(self, message):
        """Show success message to user."""
        messagebox.showinfo("Success", message)

    def show_error(self, message):
        """Show error message to user."""
        messagebox.showerror("Error", message)


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main application logic."""
    logger = setup_logging()
    logger.info("Starting Jira Project Creator")

    try:
        # Validate configuration
        Config.validate()

        # Initialize components
        ui = UserInterface()
        jira_api = JiraAPI(Config.JIRA_BASE_URL, Config.EMAIL, Config.API_TOKEN)
        project_creator = ProjectCreator()

        # Get Jira URL from user
        jira_url = ui.get_jira_url()
        if not jira_url:
            return

        # Extract ticket key
        ticket_key = jira_api.extract_ticket_key(jira_url)
        if not ticket_key:
            ui.show_error("Could not extract ticket key from URL")
            return

        # Get ticket details from Jira
        ticket = jira_api.get_ticket_info(ticket_key)
        if not ticket:
            ui.show_error("Could not retrieve ticket information from Jira")
            return

        # Get save location
        save_location = ui.get_save_location(Config.DEFAULT_SAVE_LOCATION)
        if not save_location:
            return

        # Create project folder and working file
        folder_path, file_path = project_creator.create_project(save_location, ticket, jira_url)
        if not folder_path:
            ui.show_error("Could not create project folder")
            return

        # Success!
        ui.show_success(
            f"Project created successfully!\n\n"
            f"Folder: {folder_path.name}\n"
            f"Working File: {file_path.name}"
        )
        logger.info(f"Project created: {folder_path}")

        # Ask if user wants to open the folder
        if ui.ask_open_folder():
            project_creator.open_folder(folder_path)

    except ValueError as e:
        logger.error(str(e))
        messagebox.showerror("Configuration Error", str(e))
    except Exception as e:
        error_msg = f"Unexpected error: {e}"
        logger.error(error_msg)
        messagebox.showerror("Error", error_msg)


if __name__ == "__main__":
    main()

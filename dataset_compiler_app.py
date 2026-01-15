"""
Dataset Compiler - Streamlit Application
=========================================
Compile up to 5 datasets into a single Excel workbook with one sheet per dataset.

Features:
- Jira ticket integration (auto-fetch from URL)
- Multiple data input methods: paste, CSV URL, Excel URL
- Preview and validation before export
- Formatted Excel output with hyperlinks

HOW TO RUN:
-----------
1. Ensure .env file has Jira credentials configured
2. Run: streamlit run dataset_compiler_app.py
"""

import streamlit as st
import pandas as pd
import re
from io import BytesIO
from typing import Optional, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from jira_helper import JiraClient, JiraConfig

# =============================================================================
# CONFIGURATION
# =============================================================================

MAX_DATASETS = 5
DATA_HEADER_ROW = 10  # Headers start at row 10
DATA_START_ROW = 11   # Data starts at row 11

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize a string to be a valid Excel sheet name.
    Excel rules: max 31 chars, no []:*?/\\
    """
    if not name:
        return "Sheet"
    # Remove invalid characters
    sanitized = re.sub(r'[\[\]:*?/\\]', '', name)
    # Truncate to 31 characters
    return sanitized[:31].strip() or "Sheet"


def sanitize_filename(name: str, max_length: int = 80) -> str:
    """Convert a string to a safe filename component."""
    if not name:
        return ""
    # Lowercase, replace spaces with hyphens
    sanitized = name.lower().strip()
    sanitized = re.sub(r'\s+', '-', sanitized)
    # Remove unsafe characters
    sanitized = re.sub(r'[<>:"/\\|?*\[\]]', '', sanitized)
    # Remove consecutive hyphens
    sanitized = re.sub(r'-+', '-', sanitized)
    # Truncate
    return sanitized[:max_length].strip('-')


def dedupe_sheet_names(names: List[str]) -> List[str]:
    """Ensure all sheet names are unique by appending numbers."""
    seen = {}
    result = []
    for name in names:
        base = sanitize_sheet_name(name)
        if base in seen:
            seen[base] += 1
            result.append(f"{base[:28]}_{seen[base]}")
        else:
            seen[base] = 1
            result.append(base)
    return result


def detect_delimiter(text: str) -> str:
    """Detect the most likely delimiter in pasted text."""
    # Count occurrences in first few lines
    lines = text.strip().split('\n')[:5]
    sample = '\n'.join(lines)

    counts = {
        '\t': sample.count('\t'),
        ',': sample.count(','),
        ';': sample.count(';')
    }

    # Prefer tab if present, then comma, then semicolon
    if counts['\t'] > 0:
        return '\t'
    elif counts[','] > 0:
        return ','
    elif counts[';'] > 0:
        return ';'
    return ','


def parse_pasted_data(text: str) -> pd.DataFrame:
    """Parse pasted CSV/TSV data into a DataFrame."""
    if not text or not text.strip():
        raise ValueError("No data provided")

    delimiter = detect_delimiter(text)
    try:
        df = pd.read_csv(
            BytesIO(text.encode('utf-8')),
            sep=delimiter,
            engine='python'
        )
        if df.empty:
            raise ValueError("Parsed data is empty")
        return df
    except Exception as e:
        raise ValueError(f"Failed to parse pasted data: {e}")


def parse_excel_sheet(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Parse a specific sheet from Excel file bytes."""
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
        if df.empty:
            raise ValueError(f"Sheet '{sheet_name}' is empty")
        return df
    except Exception as e:
        raise ValueError(f"Failed to parse Excel sheet: {e}")


def normalize_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize columns containing 'date' or 'month' to datetime.
    Returns a copy of the dataframe.
    """
    df = df.copy()
    for col in df.columns:
        col_lower = str(col).lower()
        if 'date' in col_lower or 'month' in col_lower:
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
            except Exception:
                pass  # Keep original if conversion fails
    return df


def get_dtype_summary(df: pd.DataFrame) -> Dict[str, str]:
    """Get a summary of column data types."""
    return {col: str(df[col].dtype) for col in df.columns}


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def create_workbook(
    jira_ticket: Dict,
    datasets: List[Dict],
    sheet_names: List[str]
) -> BytesIO:
    """
    Create an Excel workbook with formatted sheets.

    Args:
        jira_ticket: Dict with 'key', 'summary', 'url'
        datasets: List of dicts with 'df', 'source_label', 'source_url'
        sheet_names: List of sanitized, unique sheet names

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    bold_font = Font(bold=True)
    hyperlink_font = Font(color="0563C1", underline="single")
    left_align = Alignment(horizontal='left', vertical='top')

    for idx, (dataset, sheet_name) in enumerate(zip(datasets, sheet_names)):
        ws = wb.create_sheet(title=sheet_name)
        df = dataset['df']
        source_label = dataset.get('source_label', '')
        source_url = dataset.get('source_url', '')

        # --- Row 1: Jira info ---
        # A1: Jira issue key
        ws['A1'] = jira_ticket['key']
        ws['A1'].font = bold_font

        # B1: Jira summary as hyperlink
        ws['B1'] = jira_ticket['summary']
        ws['B1'].hyperlink = jira_ticket['url']
        ws['B1'].font = hyperlink_font

        # --- Row 2: Dataset source ---
        # A2: Source label
        ws['A2'] = source_label

        # B2: Source URL as hyperlink (if provided)
        if source_url:
            ws['B2'] = source_url
            ws['B2'].hyperlink = source_url
            ws['B2'].font = hyperlink_font

        # --- Row 10+: Data headers and content ---
        # Write headers at row 10
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=DATA_HEADER_ROW, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.alignment = left_align

        # Write data starting at row 11
        for row_idx, row in enumerate(df.itertuples(index=False), start=DATA_START_ROW):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Handle different types
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, pd.Timestamp):
                    cell.value = value.to_pydatetime()
                    cell.number_format = 'YYYY-MM-DD'
                else:
                    cell.value = value

        # --- Formatting ---
        # Auto-filter on header row
        if len(df.columns) > 0:
            last_col = get_column_letter(len(df.columns))
            last_row = DATA_START_ROW + len(df) - 1
            ws.auto_filter.ref = f"A{DATA_HEADER_ROW}:{last_col}{last_row}"

        # Freeze panes at A11 (row 10 and column A stay visible)
        ws.freeze_panes = 'B11'

        # Auto-width columns (approximate)
        for col_idx, col_name in enumerate(df.columns, start=1):
            # Estimate width based on header and a sample of data
            max_len = len(str(col_name))
            for value in df.iloc[:10, col_idx - 1]:
                if not pd.isna(value):
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_filename(jira_key: str, jira_summary: str) -> str:
    """Generate the output filename: WORKING - PROJ-123 - Ticket Title.xlsx"""
    # Remove unsafe characters but keep spaces
    sanitized_title = re.sub(r'[<>:"/\\|?*\[\]]', '', jira_summary)
    sanitized_title = sanitized_title.strip()[:80]
    return f"WORKING - {jira_key} - {sanitized_title}.xlsx"


# =============================================================================
# STREAMLIT UI COMPONENTS
# =============================================================================

def render_jira_section() -> Optional[Dict]:
    """Render Jira ticket input section. Returns ticket info or None."""
    st.subheader("1. Jira Ticket")

    # Check if Jira is configured
    if not JiraConfig.is_configured():
        missing = JiraConfig.get_missing()
        st.error(f"Jira not configured. Missing in .env: {', '.join(missing)}")
        return None

    col1, col2 = st.columns([3, 1])
    with col1:
        jira_url = st.text_input(
            "Jira Ticket URL",
            placeholder="https://yourcompany.atlassian.net/browse/PROJ-123",
            key="jira_url_input"
        )

    with col2:
        fetch_clicked = st.button("Fetch", use_container_width=True, disabled=not jira_url)

    # Handle fetch
    if fetch_clicked and jira_url:
        try:
            client = JiraClient()
            ticket_key = client.extract_ticket_key(jira_url)
            if not ticket_key:
                st.error("Could not extract ticket key from URL")
                return None

            with st.spinner("Fetching ticket..."):
                ticket = client.fetch_ticket(ticket_key)
                if ticket:
                    st.session_state.jira_ticket = ticket
                    st.success(f"Fetched: {ticket['key']}")
                else:
                    st.error("Could not fetch ticket")
        except ValueError as e:
            st.error(str(e))
        except PermissionError as e:
            st.error(str(e))
        except ConnectionError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"Unexpected error: {e}")

    # Display current ticket
    if 'jira_ticket' in st.session_state and st.session_state.jira_ticket:
        ticket = st.session_state.jira_ticket
        st.info(f"**{ticket['key']}** | {ticket['summary']}")
        return ticket

    return None


def render_dataset_input(dataset_idx: int) -> Optional[Dict]:
    """
    Render input controls for a single dataset.
    Returns dict with 'df', 'name', 'source_label', 'source_url' or None.
    """
    dataset_key = f"dataset_{dataset_idx}"

    with st.expander(f"Dataset {dataset_idx + 1}", expanded=(dataset_idx == 0)):
        # Source label doubles as sheet name
        col1, col2 = st.columns(2)
        with col1:
            source_label = st.text_input(
                "Data Source (also used as sheet name)",
                key=f"{dataset_key}_source_label",
                placeholder="e.g., Churn by Service"
            )
        with col2:
            source_url = st.text_input(
                "Data Source URL (optional)",
                key=f"{dataset_key}_source_url",
                placeholder="https://..."
            )

        # Input method
        input_method = st.radio(
            "Input Method",
            options=["Upload File", "Paste Data"],
            key=f"{dataset_key}_method",
            horizontal=True
        )

        df = None
        parse_error = None

        # --- Upload File (CSV or Excel) ---
        if input_method == "Upload File":
            uploaded_file = st.file_uploader(
                "Upload CSV or Excel file",
                type=["csv", "xlsx", "xls"],
                key=f"{dataset_key}_upload"
            )

            if uploaded_file:
                file_ext = uploaded_file.name.split('.')[-1].lower()

                try:
                    if file_ext == 'csv':
                        # Parse CSV directly
                        df = pd.read_csv(uploaded_file)
                        st.session_state[f"{dataset_key}_loaded_df"] = df

                    elif file_ext in ['xlsx', 'xls']:
                        # Read Excel file bytes and get sheet names
                        file_bytes = uploaded_file.read()
                        uploaded_file.seek(0)  # Reset for potential re-read

                        xl = pd.ExcelFile(BytesIO(file_bytes))
                        sheet_names = xl.sheet_names
                        st.session_state[f"{dataset_key}_excel_bytes"] = file_bytes
                        st.session_state[f"{dataset_key}_excel_sheets"] = sheet_names

                except Exception as e:
                    parse_error = str(e)

                # Sheet selector for Excel files
                if f"{dataset_key}_excel_sheets" in st.session_state:
                    sheets = st.session_state[f"{dataset_key}_excel_sheets"]
                    selected_sheet = st.selectbox(
                        "Select Sheet",
                        options=sheets,
                        key=f"{dataset_key}_selected_sheet"
                    )

                    if selected_sheet and f"{dataset_key}_excel_bytes" in st.session_state:
                        try:
                            df = parse_excel_sheet(
                                st.session_state[f"{dataset_key}_excel_bytes"],
                                selected_sheet
                            )
                        except Exception as e:
                            parse_error = str(e)

                # Use cached CSV df if available
                elif f"{dataset_key}_loaded_df" in st.session_state and file_ext == 'csv':
                    df = st.session_state[f"{dataset_key}_loaded_df"]

        # --- Paste Data ---
        elif input_method == "Paste Data":
            pasted = st.text_area(
                "Paste CSV/TSV data",
                key=f"{dataset_key}_paste",
                height=150,
                placeholder="Paste your data here (comma, tab, or semicolon separated)"
            )
            if pasted and pasted.strip():
                try:
                    df = parse_pasted_data(pasted)
                except Exception as e:
                    parse_error = str(e)

        # Show error
        if parse_error:
            st.error(f"Parse error: {parse_error}")

        # Show preview if we have data
        if df is not None and not df.empty:
            # Normalize date columns
            df = normalize_date_columns(df)

            st.markdown("**Preview**")
            st.dataframe(df.head(10), use_container_width=True)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Rows", len(df))
            with col2:
                st.metric("Columns", len(df.columns))
            with col3:
                # Check for date columns
                date_cols = [c for c in df.columns if df[c].dtype == 'datetime64[ns]']
                st.metric("Date Columns", len(date_cols))

            # Return dataset info (source_label is used as sheet name)
            return {
                'df': df,
                'name': source_label or f"Dataset {dataset_idx + 1}",
                'source_label': source_label or "",
                'source_url': source_url or ""
            }

    return None


def render_export_section(jira_ticket: Dict, datasets: List[Dict]):
    """Render the export button and download."""
    st.divider()
    st.subheader("3. Export")

    if not datasets:
        st.warning("Add at least one dataset to export")
        return

    # Generate sheet names (unique)
    raw_names = [d['name'] for d in datasets]
    sheet_names = dedupe_sheet_names(raw_names)

    # Preview what will be exported
    st.markdown("**Export Preview:**")
    preview_data = []
    for idx, (ds, sheet) in enumerate(zip(datasets, sheet_names)):
        preview_data.append({
            "Sheet": sheet,
            "Source": ds['source_label'] or "(none)",
            "Rows": len(ds['df']),
            "Columns": len(ds['df'].columns)
        })
    st.table(pd.DataFrame(preview_data))

    # Export button
    filename = generate_filename(jira_ticket['key'], jira_ticket['summary'])

    if st.button("Generate Excel", type="primary", use_container_width=True):
        try:
            with st.spinner("Generating Excel file..."):
                buffer = create_workbook(jira_ticket, datasets, sheet_names)
                st.session_state.export_buffer = buffer
                st.session_state.export_filename = filename
                st.success("Excel file generated!")
        except Exception as e:
            st.error(f"Export failed: {e}")

    # Download button
    if 'export_buffer' in st.session_state and st.session_state.export_buffer:
        st.download_button(
            label=f"Download {st.session_state.export_filename}",
            data=st.session_state.export_buffer,
            file_name=st.session_state.export_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    st.set_page_config(
        page_title="Dataset Compiler",
        page_icon="📊",
        layout="wide"
    )

    st.title("📊 Dataset Compiler")
    st.markdown("Compile up to 5 datasets into a single Excel workbook")

    # Initialize session state
    if 'jira_ticket' not in st.session_state:
        st.session_state.jira_ticket = None
    if 'export_buffer' not in st.session_state:
        st.session_state.export_buffer = None

    st.divider()

    # --- Section 1: Jira Ticket ---
    jira_ticket = render_jira_section()

    st.divider()

    # --- Section 2: Datasets ---
    st.subheader("2. Datasets")

    if not jira_ticket:
        st.info("Fetch a Jira ticket first to enable dataset input")
    else:
        datasets = []
        for idx in range(MAX_DATASETS):
            dataset = render_dataset_input(idx)
            if dataset:
                datasets.append(dataset)

        # --- Section 3: Export ---
        render_export_section(jira_ticket, datasets)


if __name__ == "__main__":
    main()

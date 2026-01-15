"""
Data Ingestion Module
=====================
Handles loading data from CSV, pasted text, and Excel files.
"""

import pandas as pd
from io import StringIO
from dataclasses import dataclass, field
from typing import Optional, Dict, Any, List
from enum import Enum


class SourceType(Enum):
    CSV = "csv"
    PASTE = "paste"
    EXCEL = "excel"


@dataclass
class DatasetMetadata:
    """Metadata about the source of a dataset."""
    source_type: SourceType
    filename: Optional[str] = None
    sheet_name: Optional[str] = None
    range_label: Optional[str] = None
    range_bounds: Optional[tuple] = None  # (min_row, max_row, min_col, max_col)


@dataclass
class Dataset:
    """
    Normalized internal representation of a dataset.
    Used throughout the pipeline.
    """
    dataset_id: int  # 1-5
    source_type: SourceType
    dataframe: pd.DataFrame
    metadata: DatasetMetadata
    detected_columns_info: Dict[str, Any] = field(default_factory=dict)
    detected_footnotes: List[str] = field(default_factory=list)
    target_sheet: Optional[str] = None  # Which template sheet this maps to

    @property
    def is_valid(self) -> bool:
        """Check if dataset has valid data."""
        return self.dataframe is not None and not self.dataframe.empty and len(self.dataframe) > 0

    @property
    def row_count(self) -> int:
        return len(self.dataframe) if self.dataframe is not None else 0

    @property
    def col_count(self) -> int:
        return len(self.dataframe.columns) if self.dataframe is not None else 0


def load_csv(file_buffer, filename: str = None) -> Optional[Dataset]:
    """
    Load data from a CSV file.

    Args:
        file_buffer: File-like object or path to CSV
        filename: Original filename for metadata

    Returns:
        Dataset object or None on error
    """
    try:
        # Try UTF-8 first, fall back to latin-1
        try:
            df = pd.read_csv(file_buffer, encoding='utf-8')
        except UnicodeDecodeError:
            if hasattr(file_buffer, 'seek'):
                file_buffer.seek(0)
            df = pd.read_csv(file_buffer, encoding='latin-1')

        metadata = DatasetMetadata(
            source_type=SourceType.CSV,
            filename=filename
        )

        return Dataset(
            dataset_id=0,  # Will be set by caller
            source_type=SourceType.CSV,
            dataframe=df,
            metadata=metadata
        )

    except Exception as e:
        print(f"Error loading CSV: {e}")
        return None


def load_pasted_data(pasted_text: str) -> Optional[Dataset]:
    """
    Parse pasted text data into a Dataset.

    Args:
        pasted_text: Tab or comma separated text data

    Returns:
        Dataset object or None on error
    """
    if not pasted_text or not pasted_text.strip():
        return None

    try:
        first_line = pasted_text.strip().split('\n')[0]
        delimiter = '\t' if '\t' in first_line else ','
        df = pd.read_csv(StringIO(pasted_text), delimiter=delimiter)

        metadata = DatasetMetadata(
            source_type=SourceType.PASTE
        )

        return Dataset(
            dataset_id=0,
            source_type=SourceType.PASTE,
            dataframe=df,
            metadata=metadata
        )

    except Exception as e:
        print(f"Error parsing pasted data: {e}")
        return None


def load_excel_range(file_buffer, sheet_name: str, min_row: int, max_row: int,
                     min_col: int, max_col: int, filename: str = None,
                     range_label: str = None) -> Optional[Dataset]:
    """
    Load data from a specific range in an Excel file.

    Args:
        file_buffer: File-like object or path to Excel file
        sheet_name: Name of the sheet to read from
        min_row, max_row, min_col, max_col: Bounds of the range
        filename: Original filename for metadata
        range_label: Human-readable label for the range

    Returns:
        Dataset object or None on error
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_buffer, read_only=True, data_only=True)
        ws = wb[sheet_name]

        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            data.append([cell.value for cell in row])

        wb.close()

        if not data:
            return None

        # First row as header
        df = pd.DataFrame(data[1:], columns=data[0])

        metadata = DatasetMetadata(
            source_type=SourceType.EXCEL,
            filename=filename,
            sheet_name=sheet_name,
            range_label=range_label,
            range_bounds=(min_row, max_row, min_col, max_col)
        )

        return Dataset(
            dataset_id=0,
            source_type=SourceType.EXCEL,
            dataframe=df,
            metadata=metadata
        )

    except Exception as e:
        print(f"Error loading Excel range: {e}")
        return None


def validate_dataset(dataset: Dataset) -> tuple[bool, Optional[str]]:
    """
    Validate a dataset for common issues.

    Returns:
        (is_valid, error_message) tuple
    """
    if dataset is None:
        return False, "No data loaded"

    df = dataset.dataframe

    if df is None:
        return False, "DataFrame is None"

    if df.empty:
        return False, "DataFrame is empty"

    if len(df) < 1:
        return False, "DataFrame has no data rows (only header)"

    if df.columns.duplicated().any():
        dupes = df.columns[df.columns.duplicated()].tolist()
        return False, f"Duplicate columns found: {dupes}"

    return True, None


def validate_dataframe(df: pd.DataFrame) -> tuple[bool, Optional[str]]:
    """
    Validate a DataFrame for common issues.
    (Kept for backward compatibility with existing code)

    Returns:
        (is_valid, error_message) tuple
    """
    if df is None:
        return False, "No data loaded"

    if df.empty:
        return False, "DataFrame is empty"

    if df.columns.duplicated().any():
        dupes = df.columns[df.columns.duplicated()].tolist()
        return False, f"Duplicate columns: {dupes}"

    return True, None

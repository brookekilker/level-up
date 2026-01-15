"""
Excel Table/Range Detection Module
==================================
Detects candidate data regions in Excel files for import.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from dataclasses import dataclass
from typing import List, Optional, Tuple
import re


@dataclass
class CandidateRange:
    """Represents a detected table-like region in an Excel sheet."""
    label: str
    sheet_name: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    range_type: str  # 'table', 'detected', 'pivot', 'manual'
    preview_df: Optional[pd.DataFrame] = None

    @property
    def bounds(self) -> Tuple[int, int, int, int]:
        return (self.min_row, self.max_row, self.min_col, self.max_col)

    @property
    def shape(self) -> Tuple[int, int]:
        return (self.max_row - self.min_row + 1, self.max_col - self.min_col + 1)

    @property
    def range_string(self) -> str:
        return f"{get_column_letter(self.min_col)}{self.min_row}:{get_column_letter(self.max_col)}{self.max_row}"

    def __str__(self):
        rows, cols = self.shape
        return f"{self.label} ({rows} rows x {cols} cols)"


def load_excel_sheets(file_path_or_buffer) -> List[str]:
    """Return list of sheet names from an Excel file."""
    wb = load_workbook(file_path_or_buffer, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def detect_defined_tables(wb, sheet_name: str) -> List[CandidateRange]:
    """Detect Excel defined tables (ListObjects) in a sheet."""
    candidates = []
    ws = wb[sheet_name]

    # Check for defined tables
    if hasattr(ws, 'tables') and ws.tables:
        for table_name, table in ws.tables.items():
            # Parse table range (e.g., "A1:H50")
            ref = table.ref
            match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', ref, re.IGNORECASE)
            if match:
                min_col = column_index_from_string(match.group(1).upper())
                min_row = int(match.group(2))
                max_col = column_index_from_string(match.group(3).upper())
                max_row = int(match.group(4))

                candidates.append(CandidateRange(
                    label=f"Table: {table_name}",
                    sheet_name=sheet_name,
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    range_type='table'
                ))

    return candidates


def detect_contiguous_ranges(wb, sheet_name: str, max_candidates: int = 5) -> List[CandidateRange]:
    """Detect contiguous rectangular data regions in a sheet."""
    ws = wb[sheet_name]
    candidates = []

    # Get all non-empty cells
    non_empty_cells = set()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5000),
                            min_col=1, max_col=min(ws.max_column, 100)):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                non_empty_cells.add((cell.row, cell.column))

    if not non_empty_cells:
        return candidates

    # Find rectangular blocks using flood-fill style detection
    visited = set()
    blocks = []

    for start_cell in sorted(non_empty_cells):
        if start_cell in visited:
            continue

        # BFS to find contiguous block
        min_row = max_row = start_cell[0]
        min_col = max_col = start_cell[1]

        # Expand to find rectangular bounds
        # Look for rows that have data in similar columns
        row_ranges = {}
        for (r, c) in non_empty_cells:
            if r not in row_ranges:
                row_ranges[r] = [c, c]
            else:
                row_ranges[r][0] = min(row_ranges[r][0], c)
                row_ranges[r][1] = max(row_ranges[r][1], c)

        # Find contiguous rows starting from this cell
        current_row = start_cell[0]
        block_rows = [current_row]

        # Expand downward
        while current_row + 1 in row_ranges:
            # Check if columns overlap significantly
            if current_row in row_ranges:
                curr_cols = set(range(row_ranges[current_row][0], row_ranges[current_row][1] + 1))
                next_cols = set(range(row_ranges[current_row + 1][0], row_ranges[current_row + 1][1] + 1))
                overlap = len(curr_cols & next_cols)
                if overlap >= min(len(curr_cols), len(next_cols)) * 0.5:
                    current_row += 1
                    block_rows.append(current_row)
                else:
                    break
            else:
                break

        if len(block_rows) >= 2:  # At least header + 1 data row
            min_row = min(block_rows)
            max_row = max(block_rows)

            # Find column bounds for this block
            cols_in_block = set()
            for r in block_rows:
                if r in row_ranges:
                    cols_in_block.update(range(row_ranges[r][0], row_ranges[r][1] + 1))

            if cols_in_block:
                min_col = min(cols_in_block)
                max_col = max(cols_in_block)

                # Mark cells as visited
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        visited.add((r, c))

                blocks.append((min_row, max_row, min_col, max_col, (max_row - min_row + 1) * (max_col - min_col + 1)))

    # Sort by size (largest first) and take top candidates
    blocks.sort(key=lambda x: x[4], reverse=True)

    for i, (min_r, max_r, min_c, max_c, _) in enumerate(blocks[:max_candidates]):
        candidates.append(CandidateRange(
            label=f"Range: {get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}",
            sheet_name=sheet_name,
            min_row=min_r,
            max_row=max_r,
            min_col=min_c,
            max_col=max_c,
            range_type='detected'
        ))

    return candidates


def detect_pivot_tables(wb, sheet_name: str) -> List[CandidateRange]:
    """Detect pivot table output regions using heuristics."""
    ws = wb[sheet_name]
    candidates = []

    pivot_keywords = ['row labels', 'column labels', 'grand total', 'sum of', 'count of',
                      'average of', 'values', '(blank)']

    # Scan for pivot table indicators
    pivot_cells = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2000),
                            min_col=1, max_col=min(ws.max_column, 50)):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(kw in cell.value.lower() for kw in pivot_keywords):
                    pivot_cells.append((cell.row, cell.column))

    if pivot_cells:
        # Find bounding box around pivot indicators
        min_row = min(c[0] for c in pivot_cells)
        max_row = max(c[0] for c in pivot_cells)
        min_col = min(c[1] for c in pivot_cells)
        max_col = max(c[1] for c in pivot_cells)

        # Expand to include adjacent data
        while min_row > 1:
            has_data = any(ws.cell(min_row - 1, c).value for c in range(min_col, max_col + 1))
            if has_data:
                min_row -= 1
            else:
                break

        while max_row < ws.max_row:
            has_data = any(ws.cell(max_row + 1, c).value for c in range(min_col, max_col + 1))
            if has_data:
                max_row += 1
            else:
                break

        candidates.append(CandidateRange(
            label=f"Pivot: {get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            sheet_name=sheet_name,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            range_type='pivot'
        ))

    return candidates


def detect_all_candidates(file_path_or_buffer, sheet_name: str) -> List[CandidateRange]:
    """
    Detect all candidate data regions in a sheet.
    Returns candidates in priority order: tables, detected ranges, pivot tables.
    """
    wb = load_workbook(file_path_or_buffer, read_only=False, data_only=True)

    candidates = []

    # 1. Defined tables (highest priority)
    candidates.extend(detect_defined_tables(wb, sheet_name))

    # 2. Contiguous detected ranges
    candidates.extend(detect_contiguous_ranges(wb, sheet_name))

    # 3. Pivot table regions
    candidates.extend(detect_pivot_tables(wb, sheet_name))

    # Remove duplicates (same bounds)
    seen_bounds = set()
    unique_candidates = []
    for c in candidates:
        if c.bounds not in seen_bounds:
            seen_bounds.add(c.bounds)
            unique_candidates.append(c)

    wb.close()
    return unique_candidates


def parse_manual_range(range_string: str) -> Optional[Tuple[int, int, int, int]]:
    """Parse a manual range string like 'A1:H50' into bounds."""
    match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_string.strip(), re.IGNORECASE)
    if match:
        min_col = column_index_from_string(match.group(1).upper())
        min_row = int(match.group(2))
        max_col = column_index_from_string(match.group(3).upper())
        max_row = int(match.group(4))
        return (min_row, max_row, min_col, max_col)
    return None


def extract_range_to_dataframe(file_path_or_buffer, sheet_name: str,
                                min_row: int, max_row: int,
                                min_col: int, max_col: int) -> pd.DataFrame:
    """Extract a specific range from an Excel sheet as a DataFrame."""
    wb = load_workbook(file_path_or_buffer, read_only=True, data_only=True)
    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        data.append([cell.value for cell in row])

    wb.close()

    if not data:
        return pd.DataFrame()

    # First row as header
    df = pd.DataFrame(data[1:], columns=data[0])
    return df


def get_range_preview(file_path_or_buffer, sheet_name: str,
                      min_row: int, max_row: int,
                      min_col: int, max_col: int,
                      preview_rows: int = 10) -> pd.DataFrame:
    """Get a preview of a range (first N rows)."""
    preview_max_row = min(min_row + preview_rows, max_row)
    return extract_range_to_dataframe(file_path_or_buffer, sheet_name,
                                       min_row, preview_max_row, min_col, max_col)

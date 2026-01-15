"""
Formatted Excel Builder - Streamlit Application
================================================
A tool to upload/paste tabular data and export it into a formatted Excel template.

Supports:
- Single dataset mode (original behavior)
- Multi-dataset mode (up to 5 datasets mapped to template sheets)
- CSV, pasted data, and Excel file inputs

HOW TO RUN:
-----------
1. Install dependencies: pip install streamlit pandas openpyxl
2. Place your template.xlsx file in the same directory as this script
3. Run: streamlit run bk_formatted_excel_app.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date, datetime
from typing import Dict, List, Optional, Any
import os

# Import local modules
from excel_extract import (
    load_excel_sheets,
    detect_all_candidates,
    extract_range_to_dataframe,
    get_range_preview,
    parse_manual_range,
    CandidateRange
)
from template_writer import TemplateWriter, SheetConfig, write_multi_dataset_template

# =============================================================================
# CONFIGURATION
# =============================================================================

TEMPLATE_PATH = "formatted_excel_template.xlsx"
DATA_START_ROW = 7
DATA_START_COL = 1
MAX_DATASETS = 5

CELL_USER_INPUT = (1, 2)      # B1 - Data Pull Name (same as filename)
CELL_SERVICES = (2, 2)        # B2 - De-duplicated services
CELL_DISTRIBUTORS = (3, 2)    # B3 - De-duplicated distributors
CELL_FOOTNOTES = (4, 2)       # B4 - Selected footnotes
CELL_DATE = (5, 2)            # B5 - Today's date

# =============================================================================
# FOOTNOTES TABLE
# =============================================================================

FOOTNOTES_TABLE = {
    "Any data before January 2023": "We recommend using 24 months historical data for the most accurate trends. With further historical data, data quality is impacted by panel changes and signal loss.",
    "Any Demographic": "Data is collected at account level, not household level.",
    "Any Demographic with projected metrics": "Data is collected at account level, not household level. Antenna does not have full coverage of panelists' demographics, but relative trends are meaningful.",
    "Any metric that includes Plan Mix, Price Paid or LTV": "Antenna does not cover 100% of plans in all cases, but relative trends are meaningful.",
    "DMA data": "Antenna does not have full coverage of panelists' zip codes, and not all zip codes map to a DMA.",
    "Churn Rate": "A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors.",
    "Survival Rate": "Survival Rate is defined as the percentage of new Subscribers in period 0 who remained Subscribed (and did not Cancel) in each period thereafter. Users who Cancel cannot re-enter the Survival Curve in subsequent months. Cohort Survival includes monthly and annual plans. A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors.",
    "M[x] Churn Rate": "M[x] Churn Rate is defined as the percentage of new Subscribers in Month 0 who Cancelled before the end of [Month x]. Users who Cancel cannot re-enter in subsequent periods. Cohort Survival includes monthly and annual plans. A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors.",
    "Average Number of Subscriptions per Subscriber": "Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Bucketed Subscriptions per Subscriber": "Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Trial Conversion Rate": "The percentage of Trials who convert to paying Subscribers. A Trial is defined as 6 months or less. Trial Conversion Rate is calculated for the conversion month. Eligible users are those whose Trial is expiring in a given month.",
    "Overlap": "Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Resubscribe Rate": "12-month Resubscribe Rate is defined as the percentage of Gross Subscriber Adds who had previously Subscribed to and since Cancelled the same Service within the prior 12 months. This metric is calculated starting 3 months after a new Service and/or Distributor launch. Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Switching Rate": "Switching Rate is defined as the percentage of users who Cancelled [Cancellation Service] and Signed-up to [Switch to Service] within 30 days. A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors. Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful. This metric is lagged by 1 month.",
    "Traders": "Traders are defined as the number of users who transitioned from one Plan or Distributor in month 0 to another Plan or Distributor in month 1 while remaining subscribed to the service. Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful. This metric is lagged by 1 month.",
    "Trading Rate": "Trading is defined as the percentage of users who transitioned from one Plan or Distributor in month 0 to another Plan or Distributor in month 1 while remaining subscribed to the service.",
    "Serial Churners – Sign-ups": "Serial Churners are users who have Canceled 3 or more Premium SVOD Subscriptions in the previous 2 years. To avoid duplication, this metric is calculated on the unique user level, not the user x Service level. Data is collected at account level, not household level.",
    "Serial Churners – Subscribers": "Serial Churners are users who have Canceled 3 or more Premium SVOD Subscriptions in the previous 2 years. To avoid duplication, this metric is calculated on the unique user level, not the user x Service level. Data is collected at account level, not household level.",
    "Promotions": "Antenna does not cover 100% of promotions in all cases, but relative trends are meaningful.",
    "Price Paid": "Data reflects standard listed prices per plan and does not include promotions or other non-standard pricing. Start and End Dates are listed only when prices change within the period.",
    "Tenure": "Subscribers have a tenure of 1 in the month they Subscribe. Tenure is calculated from the point at which Antenna's reporting began. A tenure of 48 or more months is grouped as 48+ due to differences in historical data access per distributor. Tenure is calculated at the service-distributor level. When a Subscriber switches distributors but remains subscribed to the Service, Tenure will reset to 1.",
    "Win Back Rate": "Win Back Rate is defined as the percent of Cancels in a Cancel Month that then Resubscribed to the same Service within the given number of months after the Cancel.",
    "# of Lifetimes": "A Lifetime is an uninterrupted period in which a customer remained subscribed to a service since January 2021. While Lifetimes are calculated looking back to activity since January 2021, only active Subscribers since 2023 are included in the analysis.",
    "Content Leaderboard Event": "Content releases are those which drove at least 1.5x Sign-ups compared to the previous 8-week benchmark",
    "Price Increase": "Price start dates is defined as the first date Sign-ups are seen at that price in Antenna data across distributors.",
    "Daily Cancels": "Daily Cancels do not necessarily add up to Monthly Cancels due to buyers canceling on multiple days of the month."
}

# Keywords to match in column names for auto-detection
METRIC_KEYWORDS = {
    "Churn Rate": ["churn rate", "churn_rate"],
    "Survival Rate": ["survival rate", "survival_rate"],
    "M[x] Churn Rate": ["m0 churn", "m1 churn", "m2 churn", "m3 churn", "m6 churn", "m12 churn", "month churn"],
    "Trial Conversion Rate": ["trial conversion", "trial_conversion", "conversion rate"],
    "Overlap": ["overlap"],
    "Resubscribe Rate": ["resubscribe", "resub rate"],
    "Switching Rate": ["switching rate", "switch rate"],
    "Traders": ["traders"],
    "Trading Rate": ["trading rate"],
    "Serial Churners – Sign-ups": ["serial churner", "serial_churner"],
    "Serial Churners – Subscribers": ["serial churner", "serial_churner"],
    "Promotions": ["promotion"],
    "Price Paid": ["price paid", "price_paid"],
    "Tenure": ["tenure"],
    "Win Back Rate": ["win back", "winback"],
    "# of Lifetimes": ["lifetime", "lifetimes"],
    "DMA data": ["dma"],
    "Daily Cancels": ["daily cancel"],
}

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def try_parse_number(value, preserve_percent=False):
    """Attempt to convert a value to a number.
    
    Args:
        value: The value to parse
        preserve_percent: If True, returns (number, is_percent) tuple
    """
    if pd.isna(value):
        return (None, False) if preserve_percent else None
    if isinstance(value, (int, float)):
        return (value, False) if preserve_percent else value
    if isinstance(value, str):
        stripped = value.strip()
        is_percent = stripped.endswith('%')
        cleaned = stripped.replace(',', '').replace('$', '').replace('%', '')
        try:
            num = float(cleaned)
            # If it was already a percentage string (e.g., "9.42%"), convert to decimal
            if is_percent:
                num = num / 100
            return (num, is_percent) if preserve_percent else num
        except ValueError:
            return (None, False) if preserve_percent else None
    return (None, False) if preserve_percent else None


def try_parse_date(value):
    """Attempt to parse a value as a date."""
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        date_formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%d/%m/%Y', '%d-%m-%Y',
            '%Y/%m/%d', '%b %d, %Y', '%B %d, %Y', '%d %b %Y', '%d %B %Y',
            '%m/%d/%y', '%m-%d-%y', '%Y%m%d', '%b-%y', '%b %Y', '%B %Y',
            '%m/%Y', '%m-%Y', '%Y-%m'
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def is_date_column(series):
    """Check if a column likely contains dates."""
    sample = series.dropna().head(20)
    if len(sample) == 0:
        return False
    date_count = sum(1 for v in sample if try_parse_date(v) is not None)
    return date_count / len(sample) > 0.5


def is_rate_column(col_name):
    """Check if a column represents a rate/percentage."""
    rate_keywords = ['rate', 'percent', 'pct', '%', 'ratio', 'share']
    return any(kw in str(col_name).lower() for kw in rate_keywords)


def format_column_name(col_name):
    """Format column name: remove underscores, capitalize, and hyphenate compound words."""
    formatted = str(col_name).replace('_', ' ').title()
    # Convert "Sign Ups" -> "Sign-ups", "Log Ins" -> "Log-ins", etc.
    compound_words = ['Sign Ups', 'Ad Tier']
    for word in compound_words:
        if word in formatted:
            hyphenated = word.split()[0] + '-' + word.split()[1]
            formatted = formatted.replace(word, hyphenated)
    return formatted


def format_date_for_display(value):
    """Format a date value to MMM-YY for display."""
    parsed = try_parse_date(value)
    return parsed.strftime('%b-%y') if parsed else str(value) if pd.notna(value) else ""


def format_number_for_display(value, is_rate=False):
    """Format a number for display."""
    result = try_parse_number(value, preserve_percent=True)
    num, was_percent = result if result else (None, False)
    if num is None:
        return str(value) if pd.notna(value) else ""
    if is_rate or was_percent:
        # num is already in decimal form (e.g., 0.0942 for 9.42%)
        return f"{num * 100:.2f}%"
    return f"{int(num):,}" if num == int(num) else f"{num:,.0f}"


def detect_matching_footnotes(df):
    """Auto-detect which footnotes match the data columns."""
    matched = set()
    col_names_lower = [str(col).lower() for col in df.columns]
    all_text = " ".join(col_names_lower)
    
    for metric, keywords in METRIC_KEYWORDS.items():
        for keyword in keywords:
            if keyword in all_text:
                matched.add(metric)
                break
    
    # Check for demographic columns
    demo_keywords = ['age', 'gender', 'income', 'demographic', 'ethnicity', 'race']
    if any(kw in all_text for kw in demo_keywords):
        matched.add("Any Demographic")
    
    return list(matched)


# =============================================================================
# DATA PROCESSING
# =============================================================================

def load_input(uploaded_file):
    """Load data from an uploaded CSV file."""
    try:
        try:
            return pd.read_csv(uploaded_file, encoding='utf-8')
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding='latin-1')
    except Exception as e:
        st.error(f"Error loading CSV: {e}")
        return None


def parse_pasted_data(pasted_text):
    """Parse pasted text data into a DataFrame."""
    if not pasted_text or not pasted_text.strip():
        return None
    try:
        from io import StringIO
        first_line = pasted_text.strip().split('\n')[0]
        delimiter = '\t' if '\t' in first_line else ','
        return pd.read_csv(StringIO(pasted_text), delimiter=delimiter)
    except Exception as e:
        st.error(f"Error parsing data: {e}")
        return None


def validate_dataframe(df):
    """Validate DataFrame for common issues."""
    if df is None:
        return False, "No data loaded"
    if df.empty:
        return False, "DataFrame is empty"
    if df.columns.duplicated().any():
        dupes = df.columns[df.columns.duplicated()].tolist()
        return False, f"Duplicate columns: {dupes}"
    return True, None


def convert_numeric_columns(df):
    """Convert text numbers to actual numbers in the DataFrame."""
    df_converted = df.copy()
    for col in df_converted.columns:
        if is_date_column(df_converted[col]):
            continue
        converted = df_converted[col].apply(try_parse_number)
        if converted.notna().sum() > df_converted[col].notna().sum() * 0.5:
            non_null_mask = converted.notna()
            if non_null_mask.any():
                df_converted.loc[non_null_mask, col] = converted[non_null_mask]
    return df_converted


def transform_dataframe(df):
    """Transform DataFrame for display with proper formatting."""
    df_converted = convert_numeric_columns(df)
    df_display = pd.DataFrame()
    
    for col in df.columns:
        new_col = format_column_name(col)
        if is_date_column(df[col]):
            df_display[new_col] = df[col].apply(format_date_for_display)
        elif is_rate_column(col):
            df_display[new_col] = df_converted[col].apply(lambda x: format_number_for_display(x, is_rate=True))
        else:
            num_test = df_converted[col].apply(try_parse_number)
            if num_test.notna().sum() > len(df_converted[col].dropna()) * 0.5:
                df_display[new_col] = df_converted[col].apply(format_number_for_display)
            else:
                df_display[new_col] = df_converted[col].apply(lambda x: str(x) if pd.notna(x) else "")
    return df_display


def extract_unique_values(df, column_names):
    """Extract unique values from matching columns."""
    for col in df.columns:
        if col.lower() in [n.lower() for n in column_names]:
            return ", ".join(str(v) for v in df[col].dropna().unique())
    return ""


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def write_to_template(df, template_path, data_pull_name="", selected_footnotes=None):
    """Write DataFrame to Excel template."""
    try:
        wb = load_workbook(template_path)
        ws = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.active
        
        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')
        
        # B1: Data Pull Name (same as filename component)
        ws.cell(row=CELL_USER_INPUT[0], column=CELL_USER_INPUT[1]).value = data_pull_name
        
        # B2: Services
        ws.cell(row=CELL_SERVICES[0], column=CELL_SERVICES[1]).value = extract_unique_values(df, [col for col in df.columns if "service" in col.lower()])
        
        # B3: Distributors
        ws.cell(row=CELL_DISTRIBUTORS[0], column=CELL_DISTRIBUTORS[1]).value = extract_unique_values(df, [col for col in df.columns if "distributor" in col.lower()])
        
        # B4: Selected footnotes
        if selected_footnotes:
            footnote_text = "\n\n".join([f"{FOOTNOTES_TABLE[metric]}" for metric in selected_footnotes])
            cell_b4 = ws.cell(row=CELL_FOOTNOTES[0], column=CELL_FOOTNOTES[1])
            cell_b4.value = footnote_text
            cell_b4.alignment = Alignment(wrap_text=True, vertical='top')
        
        # B5: Date
        date_cell = ws.cell(row=CELL_DATE[0], column=CELL_DATE[1])
        date_cell.value = date.today()
        date_cell.number_format = 'yyyy-mm-dd'
        
        # Convert numbers before writing
        df_converted = convert_numeric_columns(df)
        
        # Identify column types
        date_cols = [col for col in df.columns if is_date_column(df[col])]
        rate_cols = [col for col in df.columns if is_rate_column(col)]
        numeric_cols = []
        for col in df_converted.columns:
            if col not in date_cols:
                num_test = df_converted[col].apply(lambda x: try_parse_number(x, preserve_percent=True)[0] if try_parse_number(x, preserve_percent=True) else None)
                if num_test.notna().sum() > len(df_converted[col].dropna()) * 0.5:
                    numeric_cols.append(col)
        
        # Track which columns should be right-aligned (numeric/rate columns)
        right_aligned_cols = set(rate_cols) | set(numeric_cols)
        
        # Write column headers with alignment matching data alignment
        for col_idx, col_name in enumerate(df.columns):
            cell = ws.cell(row=DATA_START_ROW, column=DATA_START_COL + col_idx)
            cell.value = format_column_name(col_name)
            # Right-align headers for numeric/rate columns
            if col_name in right_aligned_cols:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        
        # Write data rows
        for row_idx, row in enumerate(df_converted.itertuples(index=False), start=1):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=DATA_START_ROW + row_idx, column=DATA_START_COL + col_idx)
                col_name = df.columns[col_idx]
                
                if col_name in date_cols:
                    parsed_date = try_parse_date(value)
                    if parsed_date:
                        cell.value = parsed_date
                        cell.number_format = 'MMM-YY'
                    else:
                        cell.value = value if pd.notna(value) else ""
                    cell.alignment = left_align
                    
                elif col_name in rate_cols:
                    result = try_parse_number(value, preserve_percent=True)
                    num, was_percent = result if result else (None, False)
                    if num is not None:
                        cell.value = num  # Already in decimal form
                        cell.number_format = '0.00%'
                    else:
                        cell.value = value if pd.notna(value) else ""
                    cell.alignment = right_align
                    
                elif col_name in numeric_cols:
                    result = try_parse_number(value, preserve_percent=True)
                    num, was_percent = result if result else (None, False)
                    if num is not None:
                        if was_percent:
                            cell.value = num  # Already in decimal form
                            cell.number_format = '0.00%'
                        else:
                            cell.value = num
                            cell.number_format = '#,##0'
                    else:
                        cell.value = value if pd.notna(value) else ""
                    cell.alignment = right_align
                    
                else:
                    cell.value = value if pd.notna(value) else ""
                    cell.alignment = left_align
        
        # Delete columns marked "Delete" (scan from right to left to preserve indices)
        cols_to_delete = []
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=DATA_START_ROW, column=col_idx).value
            if header_value and str(header_value).strip().lower() == "delete":
                cols_to_delete.append(col_idx)
        
        # Delete from right to left to avoid index shifting issues
        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except FileNotFoundError:
        st.error(f"Template not found: {template_path}")
        return None
    except Exception as e:
        st.error(f"Error writing template: {e}")
        return None


# =============================================================================
# UI COMPONENTS
# =============================================================================

def display_data_preview(df, df_transformed, max_rows=10):
    """Display data preview."""
    st.subheader("📋 Data Preview")
    st.dataframe(df_transformed.head(max_rows), use_container_width=True)
    if len(df) > max_rows:
        st.caption(f"Showing first {max_rows} of {len(df)} rows")
    
    with st.expander("🔄 Column Transformations"):
        st.dataframe(pd.DataFrame({
            'Original': df.columns.tolist(),
            'Transformed': df_transformed.columns.tolist()
        }), use_container_width=True)


def display_data_summary(df):
    """Display data summary."""
    st.subheader("📊 Data Summary")
    c1, c2, c3 = st.columns(3)
    c1.metric("Rows", len(df))
    c2.metric("Columns", len(df.columns))
    c3.metric("Cells", len(df) * len(df.columns))


# =============================================================================
# MULTI-DATASET SUPPORT
# =============================================================================

def get_template_sheets() -> List[str]:
    """Get list of sheet names from the template file."""
    try:
        wb = load_workbook(TEMPLATE_PATH, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        st.error(f"Error loading template: {e}")
        return []


def init_multi_dataset_state():
    """Initialize session state for multi-dataset mode."""
    if 'datasets' not in st.session_state:
        st.session_state.datasets = {}  # {dataset_id: {'df': df, 'footnotes': [], 'source': str, ...}}
    if 'sheet_mapping' not in st.session_state:
        st.session_state.sheet_mapping = {}  # {dataset_id: sheet_name}
    if 'template_sheets' not in st.session_state:
        st.session_state.template_sheets = get_template_sheets()
    if 'excel_files' not in st.session_state:
        st.session_state.excel_files = {}  # {dataset_id: file_bytes}
    if 'excel_candidates' not in st.session_state:
        st.session_state.excel_candidates = {}  # {dataset_id: [CandidateRange]}


def render_dataset_input(dataset_id: int) -> Optional[pd.DataFrame]:
    """
    Render input controls for a single dataset slot.
    Returns the loaded DataFrame or None.
    """
    st.markdown(f"### Dataset {dataset_id}")

    input_type = st.selectbox(
        "Input type",
        ["None", "CSV", "Paste", "Excel"],
        key=f"input_type_{dataset_id}",
        help="Select how to provide data for this dataset"
    )

    df = None

    if input_type == "CSV":
        uploaded_file = st.file_uploader(
            "Upload CSV",
            type=['csv'],
            key=f"csv_upload_{dataset_id}"
        )
        if uploaded_file:
            df = load_input(uploaded_file)
            if df is not None:
                st.success(f"Loaded {len(df)} rows, {len(df.columns)} columns")

    elif input_type == "Paste":
        pasted_data = st.text_area(
            "Paste data (tab or comma separated)",
            height=150,
            key=f"paste_data_{dataset_id}"
        )
        if st.button("Parse", key=f"parse_btn_{dataset_id}") and pasted_data:
            df = parse_pasted_data(pasted_data)
            if df is not None:
                st.session_state[f"parsed_df_{dataset_id}"] = df
                st.success(f"Parsed {len(df)} rows, {len(df.columns)} columns")
        # Retrieve previously parsed data
        if f"parsed_df_{dataset_id}" in st.session_state:
            df = st.session_state[f"parsed_df_{dataset_id}"]

    elif input_type == "Excel":
        df = render_excel_input(dataset_id)

    return df


def render_excel_input(dataset_id: int) -> Optional[pd.DataFrame]:
    """Render Excel file input with sheet and range selection."""
    uploaded_file = st.file_uploader(
        "Upload Excel file",
        type=['xlsx', 'xls'],
        key=f"excel_upload_{dataset_id}"
    )

    if not uploaded_file:
        return None

    # Cache the file bytes
    file_bytes = BytesIO(uploaded_file.read())
    uploaded_file.seek(0)

    try:
        # Get sheet names
        sheets = load_excel_sheets(file_bytes)
        file_bytes.seek(0)

        if not sheets:
            st.error("No sheets found in Excel file")
            return None

        # Sheet selection
        selected_sheet = st.selectbox(
            "Select sheet",
            sheets,
            key=f"excel_sheet_{dataset_id}"
        )

        if not selected_sheet:
            return None

        # Detect candidate ranges
        with st.spinner("Detecting data regions..."):
            file_bytes.seek(0)
            candidates = detect_all_candidates(file_bytes, selected_sheet)
            file_bytes.seek(0)

        if not candidates:
            st.warning("No data regions detected. Use manual range entry.")

        # Build selection options
        range_options = ["-- Select a range --"]
        range_map = {}

        for i, c in enumerate(candidates):
            label = f"{c.label} ({c.shape[0]} rows x {c.shape[1]} cols)"
            range_options.append(label)
            range_map[label] = c

        range_options.append("Manual range entry")

        selected_range_label = st.selectbox(
            "Select data region",
            range_options,
            key=f"excel_range_{dataset_id}"
        )

        # Handle selection
        if selected_range_label == "-- Select a range --":
            return None

        elif selected_range_label == "Manual range entry":
            manual_range = st.text_input(
                "Enter range (e.g., A1:H50)",
                key=f"manual_range_{dataset_id}",
                placeholder="A1:H50"
            )
            if manual_range:
                bounds = parse_manual_range(manual_range)
                if bounds:
                    min_row, max_row, min_col, max_col = bounds
                    file_bytes.seek(0)
                    df = extract_range_to_dataframe(
                        file_bytes, selected_sheet,
                        min_row, max_row, min_col, max_col
                    )
                    if df is not None and not df.empty:
                        st.success(f"Loaded {len(df)} rows, {len(df.columns)} columns")
                        return df
                    else:
                        st.error("No data found in specified range")
                else:
                    st.error("Invalid range format. Use format like A1:H50")
            return None

        else:
            # Selected a detected range
            candidate = range_map.get(selected_range_label)
            if candidate:
                # Show preview
                with st.expander("Preview (first 10 rows)"):
                    file_bytes.seek(0)
                    preview_df = get_range_preview(
                        file_bytes, selected_sheet,
                        candidate.min_row, candidate.max_row,
                        candidate.min_col, candidate.max_col,
                        preview_rows=10
                    )
                    if preview_df is not None:
                        st.dataframe(preview_df, use_container_width=True)

                # Load full data
                file_bytes.seek(0)
                df = extract_range_to_dataframe(
                    file_bytes, selected_sheet,
                    candidate.min_row, candidate.max_row,
                    candidate.min_col, candidate.max_col
                )
                if df is not None and not df.empty:
                    st.success(f"Loaded {len(df)} rows, {len(df.columns)} columns")
                    return df

    except Exception as e:
        st.error(f"Error processing Excel file: {e}")

    return None


def render_sheet_mapping(datasets: Dict[int, Dict]) -> Dict[int, str]:
    """
    Render sheet mapping UI for multi-dataset mode.
    Users can provide custom sheet names for each dataset.
    Dataset 1 -> Sheet 1 (first sheet), Dataset 2 -> Sheet 2 (second sheet), etc.
    Returns mapping of dataset_id -> sheet_name.
    """
    st.subheader("📑 Sheet Names")
    st.markdown("Name the sheet for each dataset. Dataset 1 will be the first sheet, Dataset 2 the second, etc.")

    mapping = {}
    used_names = set()
    errors = []

    for dataset_id, data in sorted(datasets.items()):
        if data.get('df') is None:
            continue

        df = data['df']
        source = data.get('source', 'Unknown')

        col1, col2 = st.columns([1, 2])
        with col1:
            st.markdown(f"**Dataset {dataset_id}**")
            st.caption(f"Source: {source}")
            st.caption(f"Size: {len(df)} rows x {len(df.columns)} cols")

        with col2:
            # Default sheet name based on dataset number
            default_name = f"Sheet {dataset_id}"
            sheet_name = st.text_input(
                f"Sheet name for Dataset {dataset_id}",
                value=st.session_state.get(f"sheet_name_{dataset_id}", default_name),
                key=f"sheet_name_{dataset_id}",
                placeholder=f"e.g., {default_name}",
                label_visibility="collapsed"
            )

            if sheet_name.strip():
                clean_name = sheet_name.strip()
                # Excel sheet name validation
                if len(clean_name) > 31:
                    errors.append(f"Sheet name '{clean_name}' is too long (max 31 characters)")
                elif any(c in clean_name for c in ['\\', '/', '*', '?', ':', '[', ']']):
                    errors.append(f"Sheet name '{clean_name}' contains invalid characters")
                elif clean_name.lower() in [n.lower() for n in used_names]:
                    errors.append(f"Duplicate sheet name '{clean_name}'")
                else:
                    mapping[dataset_id] = clean_name
                    used_names.add(clean_name)

    # Show errors
    for error in errors:
        st.error(error)

    return mapping if not errors else {}


def render_multi_dataset_footnotes(datasets: Dict[int, Dict]) -> Dict[int, List[str]]:
    """
    Render footnote selection for each dataset.
    Returns mapping of dataset_id -> selected_footnotes.
    """
    st.subheader("📝 Footnotes Selection")

    footnotes_map = {}

    for dataset_id, data in sorted(datasets.items()):
        df = data.get('df')
        if df is None:
            continue

        with st.expander(f"Dataset {dataset_id} Footnotes"):
            # Auto-detect footnotes
            auto_detected = detect_matching_footnotes(df)

            if auto_detected:
                st.success(f"Auto-detected {len(auto_detected)} matching footnote(s)")

            selected = st.multiselect(
                "Select footnotes:",
                options=list(FOOTNOTES_TABLE.keys()),
                default=auto_detected,
                key=f"footnotes_{dataset_id}",
                help="Footnotes will be inserted into cell B4 of the assigned sheet"
            )

            footnotes_map[dataset_id] = selected

    return footnotes_map


def write_multi_dataset_output(
    datasets: Dict[int, Dict],
    sheet_mapping: Dict[int, str],
    footnotes_map: Dict[int, List[str]],
    data_pull_name: str
) -> Optional[BytesIO]:
    """
    Write multiple datasets to template with custom sheet names.
    Dataset 1 -> first sheet, Dataset 2 -> second sheet, etc.
    Unused template sheets are deleted.
    """
    try:
        # Build dataset-sheet mapping for the writer
        # Format: {dataset_id: (sheet_name, df, footnotes, data_pull_name)}
        dataset_sheet_mapping = {}

        for dataset_id, sheet_name in sheet_mapping.items():
            data = datasets.get(dataset_id, {})
            df = data.get('df')
            if df is None:
                continue

            footnotes = footnotes_map.get(dataset_id, [])
            dataset_sheet_mapping[dataset_id] = (sheet_name, df, footnotes, data_pull_name)

        if not dataset_sheet_mapping:
            st.error("No datasets assigned to sheets")
            return None

        # Helper functions dict for the writer
        helper_funcs = {
            'convert_numeric_columns': convert_numeric_columns,
            'try_parse_number': try_parse_number,
            'try_parse_date': try_parse_date,
            'is_date_column': is_date_column,
            'is_rate_column': is_rate_column,
            'format_column_name': format_column_name,
            'extract_unique_values': extract_unique_values
        }

        # Write to template
        buffer = write_multi_dataset_template(
            TEMPLATE_PATH,
            dataset_sheet_mapping,
            FOOTNOTES_TABLE,
            helper_funcs
        )

        return buffer

    except Exception as e:
        st.error(f"Error generating output: {e}")
        import traceback
        traceback.print_exc()
        return None


def run_multi_dataset_mode():
    """Run the multi-dataset mode UI."""
    init_multi_dataset_state()

    st.markdown("---")
    st.header("📂 Data Inputs")
    st.markdown("Provide up to 5 datasets. Each can be CSV, pasted data, or Excel file.")

    # Dataset input tabs
    tabs = st.tabs([f"Dataset {i}" for i in range(1, MAX_DATASETS + 1)])

    datasets = {}

    for i, tab in enumerate(tabs, start=1):
        with tab:
            df = render_dataset_input(i)
            if df is not None:
                datasets[i] = {
                    'df': df,
                    'source': st.session_state.get(f"input_type_{i}", "Unknown"),
                    'footnotes': []
                }

    # Store datasets in session state
    st.session_state.datasets = datasets

    # Check if we have any datasets
    if not datasets:
        st.info("👆 Add at least one dataset above to continue.")
        return

    st.markdown("---")

    # Sheet naming
    mapping = render_sheet_mapping(datasets)
    st.session_state.sheet_mapping = mapping

    if not mapping:
        st.warning("⚠️ Provide a sheet name for at least one dataset to continue.")
        return

    st.markdown("---")

    # Footnotes selection
    footnotes_map = render_multi_dataset_footnotes(datasets)

    st.markdown("---")

    # Generate output
    st.subheader("📥 Generate Output")

    # Get report details from sidebar
    customer_name = st.session_state.get('customer_name', '')
    data_pull_name = st.session_state.get('data_pull_name', '')

    if not customer_name or not data_pull_name:
        st.warning("⚠️ Enter Customer Name and Data Pull Name in the sidebar to generate.")
        return

    # Validation summary
    st.markdown("**Summary:**")
    for dataset_id, sheet_name in sorted(mapping.items()):
        df = datasets[dataset_id]['df']
        st.markdown(f"- Dataset {dataset_id} ({len(df)} rows) → Sheet '{sheet_name}'")

    if st.button("🔄 Generate Multi-Dataset Excel", type="primary"):
        with st.spinner("Generating..."):
            buffer = write_multi_dataset_output(
                datasets, mapping, footnotes_map, data_pull_name
            )
            if buffer:
                output_filename = f"Antenna for {customer_name}_{data_pull_name}_{date.today().strftime('%Y%m%d')}.xlsx"
                st.session_state.multi_excel_buffer = buffer
                st.session_state.multi_output_filename = output_filename
                st.success("✅ Generated!")

    if st.session_state.get('multi_excel_buffer'):
        st.download_button(
            f"⬇️ Download: {st.session_state.multi_output_filename}",
            st.session_state.multi_excel_buffer,
            st.session_state.multi_output_filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# =============================================================================
# SINGLE DATASET MODE (ORIGINAL BEHAVIOR)
# =============================================================================

def run_single_dataset_mode():
    """Run the original single-dataset mode UI."""
    # Original single-dataset logic preserved here
    df = st.session_state.df
    if df is not None:
        is_valid, error_msg = validate_dataframe(df)
        if is_valid:
            df_transformed = transform_dataframe(df)
            display_data_preview(df, df_transformed)
            display_data_summary(df)

            # Show detected services and distributors
            services = extract_unique_values(df, [col for col in df.columns if "service" in col.lower()])
            distributors = extract_unique_values(df, [col for col in df.columns if "distributor" in col.lower()])

            if services or distributors:
                col1, col2 = st.columns(2)
                with col1:
                    if services:
                        st.info(f"**Detected Services:** {services}")
                with col2:
                    if distributors:
                        st.info(f"**Detected Distributors:** {distributors}")

            # Footnotes selection section
            st.divider()
            st.subheader("📝 Footnotes Selection (for cell B4)")

            auto_detected = st.session_state.auto_footnotes
            if auto_detected:
                st.success(f"🔍 Auto-detected {len(auto_detected)} matching footnote(s) based on your data columns.")

            # Create multiselect with auto-detected defaults
            selected_footnotes = st.multiselect(
                "Select footnotes to include:",
                options=list(FOOTNOTES_TABLE.keys()),
                default=auto_detected,
                help="These footnotes will be inserted into cell B4 of the output file."
            )

            # Show preview of selected footnotes
            if selected_footnotes:
                with st.expander(f"📖 Preview {len(selected_footnotes)} selected footnote(s)"):
                    for metric in selected_footnotes:
                        st.markdown(f"**{metric}:** {FOOTNOTES_TABLE[metric]}")
                        st.markdown("---")

            st.divider()
            st.subheader("📥 Generate Output")

            customer_name = st.session_state.get('customer_name', '')
            data_pull_name = st.session_state.get('data_pull_name', '')

            if not customer_name or not data_pull_name:
                st.warning("⚠️ Enter Customer Name and Data Pull Name to generate.")

            if st.button("🔄 Generate Excel", type="primary", disabled=not (customer_name and data_pull_name)):
                with st.spinner("Generating..."):
                    buffer = write_to_template(df, TEMPLATE_PATH, data_pull_name, selected_footnotes)
                    if buffer:
                        st.session_state.excel_buffer = buffer
                        st.session_state.output_filename = f"Antenna for {customer_name}_{data_pull_name}_{date.today().strftime('%Y%m%d')}.xlsx"
                        st.success("✅ Generated!")

            if st.session_state.get('excel_buffer'):
                st.download_button(
                    f"⬇️ Download: {st.session_state.output_filename}",
                    st.session_state.excel_buffer,
                    st.session_state.output_filename,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.error(f"❌ {error_msg}")
    else:
        st.info("👈 Upload a CSV or paste data to get started.")

        # Show footnotes reference table
        with st.expander("📚 Available Footnotes Reference"):
            footnotes_df = pd.DataFrame([
                {"Metric": k, "Footnote": v} for k, v in FOOTNOTES_TABLE.items()
            ])
            st.dataframe(footnotes_df, use_container_width=True, hide_index=True)


# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    st.set_page_config(page_title="Antenna Formatted Excel Converter", page_icon="📊", layout="wide")
    st.title("📊 Antenna - Formatted Excel Converter")
    st.markdown("Transform your regular degular data into an Antenna Style Formatted Excel that even ~ Insights would approve of!")
    st.markdown("Please leave all bugs & enhancement requests in a comment on the jira ticket [here](https://antennalive.atlassian.net/browse/DA-6973)")

    # Initialize session state
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'auto_footnotes' not in st.session_state:
        st.session_state.auto_footnotes = []

    with st.sidebar:
        st.header("⚙️ Options")

        # Mode selector
        st.subheader("📊 Mode")
        mode = st.radio(
            "Select mode:",
            ["Single Dataset", "Multi-Dataset (up to 5)"],
            help="Single Dataset: Original behavior, one dataset to one sheet. Multi-Dataset: Up to 5 datasets mapped to template sheets."
        )
        st.session_state.mode = mode

        st.divider()

        # Report Details (shared by both modes)
        st.subheader("📝 Report Details")
        customer_name = st.text_input("Customer Name *", placeholder="e.g., Netflix")
        data_pull_name = st.text_input("Data Pull Name *", placeholder="e.g., Monthly Subscribers by Plan")

        # Store in session state for multi-dataset mode
        st.session_state.customer_name = customer_name
        st.session_state.data_pull_name = data_pull_name

        if customer_name and data_pull_name:
            st.caption(f"📁 `Antenna for {customer_name}_{data_pull_name}_{date.today().strftime('%Y%m%d')}.xlsx`")

        st.divider()

        # Single dataset mode inputs in sidebar
        if mode == "Single Dataset":
            st.subheader("📂 Data Input")
            input_method = st.radio("Input method:", ["Upload CSV", "Paste Data"])

            if input_method == "Upload CSV":
                uploaded_file = st.file_uploader("Choose CSV", type=['csv'])
                if uploaded_file:
                    st.session_state.df = load_input(uploaded_file)
                    if st.session_state.df is not None:
                        st.session_state.auto_footnotes = detect_matching_footnotes(st.session_state.df)
            else:
                pasted_data = st.text_area("Paste data:", height=200)
                if st.button("Parse Data", type="primary") and pasted_data:
                    st.session_state.df = parse_pasted_data(pasted_data)
                    if st.session_state.df is not None:
                        st.session_state.auto_footnotes = detect_matching_footnotes(st.session_state.df)

    # Main content area based on mode
    if mode == "Single Dataset":
        run_single_dataset_mode()
    else:
        run_multi_dataset_mode()


if __name__ == "__main__":
    main()
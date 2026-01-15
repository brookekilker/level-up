"""
Smoke Tests for Formatted Excel Builder
=======================================
Basic verification tests for CSV/paste loading and Excel range detection.

Run with: python -m pytest tests/test_smoke.py -v
Or simply: python tests/test_smoke.py
"""

import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from io import StringIO, BytesIO
import tempfile


def test_csv_loading():
    """Test CSV loading functionality."""
    from ingestion import load_csv

    csv_data = """name,value,rate
    Product A,100,0.15
    Product B,200,0.25
    Product C,300,0.35"""

    buffer = StringIO(csv_data)
    dataset = load_csv(buffer, filename="test.csv")

    assert dataset is not None, "Dataset should not be None"
    assert dataset.dataframe is not None, "DataFrame should not be None"
    assert len(dataset.dataframe) == 3, f"Expected 3 rows, got {len(dataset.dataframe)}"
    assert len(dataset.dataframe.columns) == 3, f"Expected 3 columns, got {len(dataset.dataframe.columns)}"
    print("✓ CSV loading test passed")


def test_pasted_data_parsing():
    """Test pasted data parsing (tab-separated)."""
    from ingestion import load_pasted_data

    pasted = "col1\tcol2\tcol3\nval1\tval2\tval3\nval4\tval5\tval6"

    dataset = load_pasted_data(pasted)

    assert dataset is not None, "Dataset should not be None"
    assert dataset.dataframe is not None, "DataFrame should not be None"
    assert len(dataset.dataframe) == 2, f"Expected 2 rows, got {len(dataset.dataframe)}"
    print("✓ Pasted data parsing test passed")


def test_dataframe_validation():
    """Test DataFrame validation."""
    from ingestion import validate_dataframe

    # Valid DataFrame
    df_valid = pd.DataFrame({'a': [1, 2], 'b': [3, 4]})
    is_valid, error = validate_dataframe(df_valid)
    assert is_valid, f"Should be valid, got error: {error}"

    # Empty DataFrame
    df_empty = pd.DataFrame()
    is_valid, error = validate_dataframe(df_empty)
    assert not is_valid, "Empty DataFrame should be invalid"

    # Duplicate columns
    df_dupes = pd.DataFrame([[1, 2, 3]], columns=['a', 'b', 'a'])
    is_valid, error = validate_dataframe(df_dupes)
    assert not is_valid, "DataFrame with duplicate columns should be invalid"

    print("✓ DataFrame validation test passed")


def test_excel_range_detection():
    """Test Excel range detection functionality."""
    from openpyxl import Workbook
    from excel_extract import detect_all_candidates, extract_range_to_dataframe

    # Create a test Excel file in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Write some test data
    test_data = [
        ["Header1", "Header2", "Header3"],
        ["A1", "B1", "C1"],
        ["A2", "B2", "C2"],
        ["A3", "B3", "C3"],
    ]
    for row_idx, row in enumerate(test_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Test candidate detection
    candidates = detect_all_candidates(buffer, "TestSheet")
    assert len(candidates) > 0, "Should detect at least one candidate range"

    # Test range extraction
    buffer.seek(0)
    df = extract_range_to_dataframe(buffer, "TestSheet", 1, 4, 1, 3)
    assert df is not None, "DataFrame should not be None"
    assert len(df) == 3, f"Expected 3 data rows, got {len(df)}"
    assert len(df.columns) == 3, f"Expected 3 columns, got {len(df.columns)}"

    print("✓ Excel range detection test passed")


def test_manual_range_parsing():
    """Test manual range string parsing."""
    from excel_extract import parse_manual_range

    # Valid ranges
    bounds = parse_manual_range("A1:H50")
    assert bounds == (1, 50, 1, 8), f"Expected (1, 50, 1, 8), got {bounds}"

    bounds = parse_manual_range("B2:D10")
    assert bounds == (2, 10, 2, 4), f"Expected (2, 10, 2, 4), got {bounds}"

    # Invalid range
    bounds = parse_manual_range("invalid")
    assert bounds is None, "Invalid range should return None"

    print("✓ Manual range parsing test passed")


def test_template_writer_basics():
    """Test basic template writer functionality."""
    from openpyxl import Workbook
    from template_writer import TemplateWriter, SheetConfig

    # Create a test template
    wb = Workbook()
    for i in range(1, 6):
        if i == 1:
            ws = wb.active
            ws.title = f"Sheet{i}"
        else:
            wb.create_sheet(f"Sheet{i}")

    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name
        wb.save(temp_path)

    try:
        # Test loading template
        writer = TemplateWriter(temp_path)
        sheets = writer.load_template()

        assert len(sheets) == 5, f"Expected 5 sheets, got {len(sheets)}"
        assert "Sheet1" in sheets, "Sheet1 should exist"

        # Test sheet deletion
        writer.delete_unused_sheets(["Sheet1", "Sheet3"])

        # Save and verify
        buffer = writer.save_to_buffer()
        assert buffer is not None, "Buffer should not be None"

        # Reload and check sheets
        from openpyxl import load_workbook
        buffer.seek(0)
        wb_check = load_workbook(buffer)
        remaining_sheets = wb_check.sheetnames
        assert len(remaining_sheets) == 2, f"Expected 2 sheets after deletion, got {len(remaining_sheets)}"
        assert "Sheet1" in remaining_sheets, "Sheet1 should remain"
        assert "Sheet3" in remaining_sheets, "Sheet3 should remain"

        writer.close()
        print("✓ Template writer basics test passed")

    finally:
        os.unlink(temp_path)


def test_helper_functions():
    """Test helper functions from main app."""
    # Import main app functions
    import bk_formatted_excel_app as app

    # Test number parsing
    result = app.try_parse_number("1,234")
    assert result == 1234, f"Expected 1234, got {result}"

    result = app.try_parse_number("9.42%", preserve_percent=True)
    assert result[0] is not None, "Should parse percentage"
    assert result[1] is True, "Should indicate percentage"

    # Test date column detection
    df = pd.DataFrame({'date': ['2023-01-01', '2023-02-01', '2023-03-01']})
    assert app.is_date_column(df['date']), "Should detect date column"

    df_not_date = pd.DataFrame({'values': [100, 200, 300]})
    assert not app.is_date_column(df_not_date['values']), "Should not detect non-date column as date"

    # Test rate column detection
    assert app.is_rate_column("churn_rate"), "Should detect rate column"
    assert app.is_rate_column("Conversion Rate"), "Should detect rate column"
    assert not app.is_rate_column("subscribers"), "Should not detect non-rate column"

    # Test column name formatting
    assert app.format_column_name("some_column_name") == "Some Column Name"
    assert "Sign-ups" in app.format_column_name("sign_ups") or "Sign Ups" in app.format_column_name("sign_ups")

    # Test footnote detection
    df_with_churn = pd.DataFrame({'churn_rate': [0.1, 0.2], 'subscribers': [100, 200]})
    footnotes = app.detect_matching_footnotes(df_with_churn)
    assert "Churn Rate" in footnotes, "Should detect Churn Rate footnote"

    print("✓ Helper functions test passed")


def run_all_tests():
    """Run all smoke tests."""
    print("\n" + "=" * 50)
    print("Running Smoke Tests for Formatted Excel Builder")
    print("=" * 50 + "\n")

    tests = [
        test_csv_loading,
        test_pasted_data_parsing,
        test_dataframe_validation,
        test_excel_range_detection,
        test_manual_range_parsing,
        test_template_writer_basics,
        test_helper_functions,
    ]

    passed = 0
    failed = 0

    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"✗ {test.__name__} FAILED: {e}")
            failed += 1
        except Exception as e:
            print(f"✗ {test.__name__} ERROR: {e}")
            failed += 1

    print("\n" + "=" * 50)
    print(f"Results: {passed} passed, {failed} failed")
    print("=" * 50)

    return failed == 0


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
